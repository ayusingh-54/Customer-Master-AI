"""
Excel Export Agent
==================
Creates rich, multi-sheet Excel workbooks from the customer master data.
All exports go to the user's Desktop by default (or a custom path).
"""

import os
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from demo_db import get_connection

# ── Colour palette ─────────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MID_BLUE    = "2E75B6"
LIGHT_BLUE  = "BDD7EE"
VERY_LIGHT  = "DEEAF1"
WHITE       = "FFFFFF"
ORANGE      = "F4B942"
RED         = "C00000"
GREEN       = "70AD47"
YELLOW      = "FFD966"
GREY        = "595959"
LIGHT_GREY  = "F2F2F2"

LIFECYCLE_COLOURS = {
    "ACTIVE":   GREEN,
    "PROSPECT": MID_BLUE,
    "AT-RISK":  ORANGE,
    "DORMANT":  YELLOW,
    "INACTIVE": RED,
}

OUTPUT_DIR = os.path.join(os.path.expanduser("~"), "Desktop")


def _get_output_path(filename: str, output_dir: str = None) -> str:
    folder = output_dir or OUTPUT_DIR
    os.makedirs(folder, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base, ext = os.path.splitext(filename)
    return os.path.join(folder, f"{base}_{ts}{ext or '.xlsx'}")


# ── Style helpers ──────────────────────────────────────────────────────────────

def _header_font(bold=True, size=11, colour=WHITE):
    return Font(name="Calibri", bold=bold, size=size, color=colour)

def _body_font(bold=False, size=10, colour="000000"):
    return Font(name="Calibri", bold=bold, size=size, color=colour)

def _fill(hex_colour):
    return PatternFill("solid", fgColor=hex_colour)

def _border():
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def _apply_header_row(ws, headers: list, row: int = 1,
                      bg: str = DARK_BLUE, fg: str = WHITE,
                      height: float = 28):
    ws.row_dimensions[row].height = height
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font      = _header_font(colour=fg)
        cell.fill      = _fill(bg)
        cell.alignment = _center()
        cell.border    = _border()

def _auto_width(ws, min_w=8, max_w=50):
    for col in ws.columns:
        best = min_w
        for cell in col:
            if cell.value:
                best = min(max(best, len(str(cell.value)) + 2), max_w)
        ws.column_dimensions[get_column_letter(col[0].column)].width = best

def _title_row(ws, title: str, subtitle: str, ncols: int):
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 22
    t = ws.cell(row=1, column=1, value=title)
    t.font      = Font(name="Calibri", bold=True, size=16, color=WHITE)
    t.fill      = _fill(DARK_BLUE)
    t.alignment = _left()
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)

    s = ws.cell(row=2, column=1, value=subtitle)
    s.font      = Font(name="Calibri", italic=True, size=10, color=GREY)
    s.fill      = _fill(LIGHT_BLUE)
    s.alignment = _left()
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT 1 — Full Customer Master Report
# ══════════════════════════════════════════════════════════════════════════════

def export_customers_master(output_dir: str = None) -> dict:
    """
    Full customer master report with 4 sheets:
      1. Customer Accounts (with credit utilisation)
      2. Party Sites (addresses)
      3. Contact Points
      4. Lifecycle Summary (with chart)
    """
    conn = get_connection()
    c = conn.cursor()

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Sheet 1: Customer Accounts ────────────────────────────────────────────
    ws1 = wb.create_sheet("Customer Accounts")
    headers1 = [
        "Account ID", "Account #", "Party Name", "Party Type",
        "Credit Limit ($)", "Open Balance ($)", "Utilisation (%)",
        "Avg Days to Pay", "Last Order Date", "Lifecycle State",
        "On Hold", "Status",
    ]
    _title_row(ws1, "Customer Accounts — Master Report",
               f"Generated: {datetime.now():%Y-%m-%d %H:%M}  |  Source: Oracle EBS TCA",
               len(headers1))
    _apply_header_row(ws1, headers1, row=3, bg=MID_BLUE)

    accounts = c.execute("""
        SELECT ca.cust_account_id, ca.account_number, hp.party_name, hp.party_type,
               ca.credit_limit,
               COALESCE((SELECT SUM(ps.amount_remaining)
                         FROM ar_payment_schedules ps
                         WHERE ps.cust_account_id = ca.cust_account_id
                           AND ps.status = 'OP'), 0) AS open_balance,
               ca.avg_days_to_pay, ca.last_order_date,
               ca.lifecycle_state, ca.on_hold, ca.status
        FROM hz_cust_accounts ca
        JOIN hz_parties hp ON hp.party_id = ca.party_id
        ORDER BY ca.lifecycle_state, ca.credit_limit DESC
    """).fetchall()

    for i, row in enumerate(accounts, 4):
        aid, anum, pname, ptype, climit, obal, adtp, lod, lstate, onhold, status = row
        util = (obal / climit * 100) if climit else 0
        util_str = f"{util:.1f}%"

        vals = [aid, anum, pname, ptype,
                round(climit, 2), round(obal, 2), util_str,
                round(adtp or 0, 1), lod, lstate,
                "YES" if onhold else "NO", status]

        bg = LIGHT_GREY if i % 2 == 0 else WHITE
        for col, val in enumerate(vals, 1):
            cell = ws1.cell(row=i, column=col, value=val)
            cell.font      = _body_font()
            cell.border    = _border()
            cell.alignment = _left()
            cell.fill      = _fill(bg)

        # Colour lifecycle state cell
        lc = LIFECYCLE_COLOURS.get(lstate, "FFFFFF")
        ws1.cell(row=i, column=10).fill = _fill(lc)
        ws1.cell(row=i, column=10).font = Font(name="Calibri", bold=True,
                                               size=10, color=WHITE if lstate in ("ACTIVE","INACTIVE") else "000000")
        # Colour utilisation red if >80%
        if util >= 80:
            ws1.cell(row=i, column=7).fill = _fill("FFE0E0")
            ws1.cell(row=i, column=7).font = Font(name="Calibri", bold=True, size=10, color=RED)

        # On hold — red
        if onhold:
            ws1.cell(row=i, column=11).fill = _fill("FFE0E0")
            ws1.cell(row=i, column=11).font = Font(name="Calibri", bold=True, size=10, color=RED)

    _auto_width(ws1)
    ws1.freeze_panes = "A4"

    # ── Sheet 2: Party Sites ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Party Sites")
    headers2 = ["Site ID", "Party Name", "Address", "City", "State",
                 "Postal Code", "Country", "Validated", "Latitude", "Longitude"]
    _title_row(ws2, "Party Sites — Addresses",
               f"Generated: {datetime.now():%Y-%m-%d %H:%M}", len(headers2))
    _apply_header_row(ws2, headers2, row=3, bg=MID_BLUE)

    sites = c.execute("""
        SELECT ps.party_site_id, hp.party_name,
               ps.address_line1, ps.city, ps.state, ps.postal_code, ps.country,
               ps.validated, ps.lat, ps.lon
        FROM hz_party_sites ps
        JOIN hz_parties hp ON hp.party_id = ps.party_id
        ORDER BY hp.party_name
    """).fetchall()

    for i, row in enumerate(sites, 4):
        bg = LIGHT_GREY if i % 2 == 0 else WHITE
        for col, val in enumerate(row, 1):
            cell = ws2.cell(row=i, column=col, value=val)
            cell.font = _body_font(); cell.border = _border(); cell.alignment = _left()
            cell.fill = _fill(bg)
        # Colour validated
        val_cell = ws2.cell(row=i, column=8)
        if row[7] == 1:
            val_cell.value = "YES"
            val_cell.fill = _fill("E2EFDA"); val_cell.font = Font(name="Calibri", bold=True, color=GREEN, size=10)
        else:
            val_cell.value = "NO"
            val_cell.fill = _fill("FFE0E0"); val_cell.font = Font(name="Calibri", bold=True, color=RED, size=10)
    _auto_width(ws2)
    ws2.freeze_panes = "A4"

    # ── Sheet 3: Contact Points ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Contact Points")
    headers3 = ["Contact ID", "Party Name", "Type", "Value", "Status"]
    _title_row(ws3, "Contact Points",
               f"Generated: {datetime.now():%Y-%m-%d %H:%M}", len(headers3))
    _apply_header_row(ws3, headers3, row=3, bg=MID_BLUE)

    contacts = c.execute("""
        SELECT cp.contact_point_id, hp.party_name, cp.contact_type, cp.contact_value, cp.status
        FROM hz_contact_points cp
        JOIN hz_parties hp ON hp.party_id = cp.party_id
        ORDER BY hp.party_name, cp.contact_type
    """).fetchall()

    for i, row in enumerate(contacts, 4):
        bg = LIGHT_GREY if i % 2 == 0 else WHITE
        for col, val in enumerate(row, 1):
            cell = ws3.cell(row=i, column=col, value=val)
            cell.font = _body_font(); cell.border = _border(); cell.alignment = _left()
            cell.fill = _fill(bg)
        status_cell = ws3.cell(row=i, column=5)
        if row[4] == "I":
            status_cell.fill = _fill("FFE0E0")
            status_cell.font = Font(name="Calibri", bold=True, color=RED, size=10)
            status_cell.value = "INVALID"
    _auto_width(ws3)
    ws3.freeze_panes = "A4"

    # ── Sheet 4: Lifecycle Summary + Chart ────────────────────────────────────
    ws4 = wb.create_sheet("Lifecycle Summary")
    _title_row(ws4, "Customer Lifecycle Summary",
               f"Generated: {datetime.now():%Y-%m-%d %H:%M}", 4)

    lifecycle_counts = c.execute("""
        SELECT lifecycle_state, COUNT(*) as cnt
        FROM hz_cust_accounts
        GROUP BY lifecycle_state
        ORDER BY CASE lifecycle_state
            WHEN 'ACTIVE'   THEN 1
            WHEN 'PROSPECT' THEN 2
            WHEN 'AT-RISK'  THEN 3
            WHEN 'DORMANT'  THEN 4
            WHEN 'INACTIVE' THEN 5
            ELSE 6 END
    """).fetchall()

    headers4 = ["Lifecycle State", "Account Count", "% of Total", "Action Required"]
    _apply_header_row(ws4, headers4, row=3, bg=MID_BLUE)

    total = sum(r[1] for r in lifecycle_counts)
    action_map = {
        "ACTIVE":   "Monitor credit, update addresses",
        "PROSPECT": "Enrich contact data, link to CRM",
        "AT-RISK":  "Alert account manager, analyse churn",
        "DORMANT":  "Auto-reduce credit, send win-back",
        "INACTIVE": "Archive, retain 7 years (compliance)",
    }
    for i, (state, cnt) in enumerate(lifecycle_counts, 4):
        pct = f"{cnt/total*100:.1f}%" if total else "0%"
        bg = LIFECYCLE_COLOURS.get(state, WHITE)
        vals = [state, cnt, pct, action_map.get(state, "")]
        for col, val in enumerate(vals, 1):
            cell = ws4.cell(row=i, column=col, value=val)
            cell.border    = _border()
            cell.alignment = _left()
            if col == 1:
                cell.fill = _fill(bg)
                cell.font = Font(name="Calibri", bold=True, size=10,
                                 color=WHITE if state in ("ACTIVE","INACTIVE") else "000000")
            else:
                cell.font = _body_font()
                cell.fill = _fill(LIGHT_GREY if i % 2 == 0 else WHITE)

    # Bar chart
    data = Reference(ws4, min_col=2, min_row=3, max_row=3 + len(lifecycle_counts))
    cats = Reference(ws4, min_col=1, min_row=4, max_row=3 + len(lifecycle_counts))
    chart = BarChart()
    chart.type  = "col"
    chart.title = "Accounts by Lifecycle State"
    chart.y_axis.title = "Account Count"
    chart.x_axis.title = "State"
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width  = 18
    chart.height = 12
    chart.series[0].graphicalProperties.solidFill = MID_BLUE
    ws4.add_chart(chart, "F3")

    _auto_width(ws4)

    conn.close()

    path = _get_output_path("Customer_Master_Report", output_dir)
    wb.save(path)

    return {
        "status":  "SUCCESS",
        "file":    path,
        "sheets":  ["Customer Accounts", "Party Sites", "Contact Points", "Lifecycle Summary"],
        "rows": {
            "accounts": len(accounts),
            "sites":    len(sites),
            "contacts": len(contacts),
        },
    }


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT 2 — Credit Risk Report
# ══════════════════════════════════════════════════════════════════════════════

def export_credit_report(output_dir: str = None) -> dict:
    """
    Credit risk report with payment performance scoring and recommendations.
    """
    import sys, os
    sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
    from agents.credit import evaluate_credit

    conn = get_connection()
    c = conn.cursor()
    ids = [r[0] for r in c.execute(
        "SELECT cust_account_id FROM hz_cust_accounts WHERE status='A'"
    ).fetchall()]
    conn.close()

    evaluations = [evaluate_credit(aid) for aid in ids]
    evaluations.sort(key=lambda x: x.get("score", 0))  # worst first

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Credit Risk Report")

    headers = [
        "Account ID", "Account #", "Party Name",
        "Current Limit ($)", "Open Balance ($)", "Utilisation",
        "Avg Days to Pay", "Overdue Invoices", "Return Rate",
        "Score", "Recommendation", "New Limit ($)", "Change", "Key Reason",
    ]
    _title_row(ws, "Credit Risk Report — Auto-Adjustment Analysis",
               f"Generated: {datetime.now():%Y-%m-%d %H:%M}  |  Scoring: payment speed + overdue + returns",
               len(headers))
    _apply_header_row(ws, headers, row=3, bg=DARK_BLUE)

    REC_COLOURS = {
        "INCREASE":  GREEN,
        "NO_CHANGE": MID_BLUE,
        "DECREASE":  ORANGE,
        "HOLD":      RED,
    }

    for i, ev in enumerate(evaluations, 4):
        if "error" in ev:
            continue
        change = ""
        if ev["recommendation"] == "INCREASE":
            change = f"+{ev['change_pct']}"
        elif ev["recommendation"] in ("DECREASE", "HOLD"):
            change = f"-{ev['change_pct']}"

        vals = [
            ev["cust_account_id"], ev["account_number"], ev["party_name"],
            ev["current_limit"], ev["open_balance"], ev["credit_utilisation"],
            ev["avg_days_to_pay"], ev["overdue_invoices"], ev["return_rate"],
            ev["score"], ev["recommendation"], ev["new_limit"], change,
            ev["reasons"][0] if ev["reasons"] else "",
        ]
        bg = LIGHT_GREY if i % 2 == 0 else WHITE
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = _body_font(); cell.border = _border(); cell.alignment = _left()
            cell.fill = _fill(bg)

        # Colour recommendation
        rec_col = ws.cell(row=i, column=11)
        rec = ev["recommendation"]
        rec_col.fill = _fill(REC_COLOURS.get(rec, WHITE))
        rec_col.font = Font(name="Calibri", bold=True, size=10,
                            color=WHITE if rec in ("INCREASE","HOLD") else "000000")

        # Score heat
        score = ev["score"]
        score_cell = ws.cell(row=i, column=10)
        if score >= 2:
            score_cell.fill = _fill(GREEN); score_cell.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
        elif score <= -2:
            score_cell.fill = _fill(RED);   score_cell.font = Font(name="Calibri", bold=True, color=WHITE, size=10)

    _auto_width(ws)
    ws.freeze_panes = "A4"

    path = _get_output_path("Credit_Risk_Report", output_dir)
    wb.save(path)

    return {
        "status":   "SUCCESS",
        "file":     path,
        "accounts": len(evaluations),
        "increase": sum(1 for e in evaluations if e.get("recommendation") == "INCREASE"),
        "decrease": sum(1 for e in evaluations if e.get("recommendation") == "DECREASE"),
        "hold":     sum(1 for e in evaluations if e.get("recommendation") == "HOLD"),
        "no_change":sum(1 for e in evaluations if e.get("recommendation") == "NO_CHANGE"),
    }


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT 3 — Duplicate Parties Report
# ══════════════════════════════════════════════════════════════════════════════

def export_duplicates_report(threshold: float = 0.88, output_dir: str = None) -> dict:
    """
    Export potential duplicate parties to Excel for review and merge decisions.
    """
    from agents.deduplication import find_duplicates
    result = find_duplicates(threshold=threshold)

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Duplicate Parties")

    headers = [
        "Group #", "Golden Party ID", "Golden Party Name",
        "Duplicate Party ID", "Duplicate Party Name",
        "Similarity Score", "Exact Key Match", "Action",
    ]
    _title_row(ws, "Duplicate Parties — Deduplication Report",
               f"Generated: {datetime.now():%Y-%m-%d %H:%M}  |  Threshold: {threshold:.0%}  |  Groups: {result['total_duplicate_groups']}",
               len(headers))
    _apply_header_row(ws, headers, row=3, bg=DARK_BLUE)

    row_num = 4
    for g_num, group in enumerate(result["groups"], 1):
        golden = group["golden_candidate"]
        for dup in group["duplicates"]:
            sim = dup["similarity"]
            action = "AUTO-MERGE" if sim >= 0.95 or dup["exact_key"] else "REVIEW"
            vals = [
                g_num, golden["party_id"], golden["party_name"],
                dup["party_id"], dup["party_name"],
                f"{sim:.1%}", "YES" if dup["exact_key"] else "NO", action,
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.font = _body_font(); cell.border = _border(); cell.alignment = _left()
                cell.fill = _fill(LIGHT_GREY if row_num % 2 == 0 else WHITE)

            # Action colour
            act_cell = ws.cell(row=row_num, column=8)
            if action == "AUTO-MERGE":
                act_cell.fill = _fill(GREEN); act_cell.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
            else:
                act_cell.fill = _fill(ORANGE); act_cell.font = Font(name="Calibri", bold=True, size=10)

            # Similarity colour
            sim_cell = ws.cell(row=row_num, column=6)
            if sim >= 0.95:
                sim_cell.fill = _fill(GREEN); sim_cell.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
            elif sim >= 0.88:
                sim_cell.fill = _fill(YELLOW); sim_cell.font = Font(name="Calibri", bold=True, size=10)
            row_num += 1

    _auto_width(ws)
    ws.freeze_panes = "A4"

    path = _get_output_path("Duplicate_Parties_Report", output_dir)
    wb.save(path)

    return {
        "status":       "SUCCESS",
        "file":         path,
        "groups_found": result["total_duplicate_groups"],
        "rows_written": row_num - 4,
    }


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT 4 — Audit Log
# ══════════════════════════════════════════════════════════════════════════════

def export_audit_log(workflow: str = None, output_dir: str = None) -> dict:
    """
    Export the full audit log to Excel.
    """
    conn = get_connection()
    c = conn.cursor()
    if workflow:
        rows = c.execute(
            "SELECT * FROM audit_log WHERE workflow=? ORDER BY log_id DESC", (workflow,)
        ).fetchall()
    else:
        rows = c.execute("SELECT * FROM audit_log ORDER BY log_id DESC").fetchall()
    conn.close()

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Audit Log")

    wf_title = f"  |  Workflow: {workflow}" if workflow else ""
    headers = ["Log ID", "Workflow", "Entity Type", "Entity ID", "Action", "Details", "Performed At"]
    _title_row(ws, f"AI Audit Log{wf_title}",
               f"Generated: {datetime.now():%Y-%m-%d %H:%M}  |  Total entries: {len(rows)}",
               len(headers))
    _apply_header_row(ws, headers, row=3, bg=DARK_BLUE)

    WORKFLOW_COLOURS = {
        "DEDUPLICATION":    MID_BLUE,
        "ADDRESS_VALIDATION": GREEN,
        "CREDIT_ADJUSTMENT": ORANGE,
        "CONTACT_MAINTENANCE": "7030A0",
        "RELATIONSHIP_MGMT":  "00B0F0",
        "ARCHIVING":          RED,
    }

    for i, row in enumerate(rows, 4):
        r = dict(row)
        bg = LIGHT_GREY if i % 2 == 0 else WHITE
        vals = [r["log_id"], r["workflow"], r["entity_type"],
                r["entity_id"], r["action"], r["details"], r["performed_at"]]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = _body_font(size=9); cell.border = _border()
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=(col == 6))
            cell.fill = _fill(bg)

        wf_cell = ws.cell(row=i, column=2)
        wf_col = WORKFLOW_COLOURS.get(r["workflow"], MID_BLUE)
        wf_cell.fill = _fill(wf_col)
        wf_cell.font = Font(name="Calibri", bold=True, size=9, color=WHITE)

    _auto_width(ws)
    ws.column_dimensions["F"].width = 60  # Details column
    ws.freeze_panes = "A4"

    path = _get_output_path(
        f"Audit_Log{'_' + workflow if workflow else ''}",
        output_dir
    )
    wb.save(path)

    return {
        "status":  "SUCCESS",
        "file":    path,
        "entries": len(rows),
        "workflow_filter": workflow,
    }


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT 5 — Dormant Accounts Report
# ══════════════════════════════════════════════════════════════════════════════

def export_dormant_report(output_dir: str = None) -> dict:
    """
    Export dormant/at-risk accounts eligible for archiving review.
    """
    from agents.archiving import scan_dormant_accounts
    result = scan_dormant_accounts(dry_run=True)

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Dormant Accounts")

    headers = [
        "Account ID", "Account #", "Party Name",
        "Last Order Date", "Days Inactive", "Can Archive", "Blockers",
    ]
    _title_row(ws, "Dormant Account Review Report",
               f"Generated: {datetime.now():%Y-%m-%d %H:%M}  |  "
               f"Found: {result['total_found']}  |  Archivable: {result['can_archive']}  |  Blocked: {result['blocked']}",
               len(headers))
    _apply_header_row(ws, headers, row=3, bg=DARK_BLUE)

    for i, acct in enumerate(result["accounts"], 4):
        bg = LIGHT_GREY if i % 2 == 0 else WHITE
        vals = [
            acct["cust_account_id"], acct["account_number"], acct["party_name"],
            acct["last_order_date"], acct["days_inactive"],
            "YES" if acct["can_archive"] else "NO",
            "; ".join(acct["blockers"]) if acct["blockers"] else "",
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = _body_font(); cell.border = _border(); cell.alignment = _left()
            cell.fill = _fill(bg)

        arch_cell = ws.cell(row=i, column=6)
        if acct["can_archive"]:
            arch_cell.fill = _fill(GREEN); arch_cell.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
        else:
            arch_cell.fill = _fill(ORANGE); arch_cell.font = Font(name="Calibri", bold=True, size=10)

    _auto_width(ws)
    ws.freeze_panes = "A4"

    path = _get_output_path("Dormant_Accounts_Report", output_dir)
    wb.save(path)

    return {
        "status":      "SUCCESS",
        "file":        path,
        "total_found": result["total_found"],
        "can_archive": result["can_archive"],
        "blocked":     result["blocked"],
    }


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT 6 — Full Dashboard (all-in-one workbook)
# ══════════════════════════════════════════════════════════════════════════════

def export_full_dashboard(output_dir: str = None) -> dict:
    """
    One workbook with ALL reports:
    Overview + Accounts + Sites + Contacts + Credit + Duplicates + Dormant + Audit
    """
    conn = get_connection()
    c = conn.cursor()

    wb = Workbook()
    wb.remove(wb.active)

    # ── Overview / KPI sheet ──────────────────────────────────────────────────
    ws0 = wb.create_sheet("Overview", 0)
    _title_row(ws0, "Customer Master Data AI — Executive Dashboard",
               f"Generated: {datetime.now():%Y-%m-%d %H:%M}  |  Source: Oracle EBS TCA  |  Mode: DEMO",
               6)

    kpis = [
        ("Total Parties",           c.execute("SELECT COUNT(*) FROM hz_parties WHERE status='A'").fetchone()[0]),
        ("Customer Accounts",       c.execute("SELECT COUNT(*) FROM hz_cust_accounts WHERE status='A'").fetchone()[0]),
        ("Party Sites",             c.execute("SELECT COUNT(*) FROM hz_party_sites").fetchone()[0]),
        ("Active Contacts",         c.execute("SELECT COUNT(*) FROM hz_contact_points WHERE status='A'").fetchone()[0]),
        ("Total Relationships",     c.execute("SELECT COUNT(*) FROM hz_relationships WHERE status='A'").fetchone()[0]),
        ("Unvalidated Addresses",   c.execute("SELECT COUNT(*) FROM hz_party_sites WHERE validated=0").fetchone()[0]),
        ("Open AR Balance",         c.execute("SELECT COALESCE(SUM(amount_remaining),0) FROM ar_payment_schedules WHERE status='OP'").fetchone()[0]),
        ("Accounts On Hold",        c.execute("SELECT COUNT(*) FROM hz_cust_accounts WHERE on_hold=1").fetchone()[0]),
    ]

    ws0.row_dimensions[4].height = 22
    ws0.cell(row=4, column=1, value="KPI").font = _header_font(colour=DARK_BLUE)
    ws0.cell(row=4, column=1).fill = _fill(LIGHT_BLUE)
    ws0.cell(row=4, column=2, value="Value").font = _header_font(colour=DARK_BLUE)
    ws0.cell(row=4, column=2).fill = _fill(LIGHT_BLUE)

    for i, (kpi, val) in enumerate(kpis, 5):
        ws0.cell(row=i, column=1, value=kpi).font  = _body_font(bold=True)
        ws0.cell(row=i, column=1).fill  = _fill(VERY_LIGHT)
        ws0.cell(row=i, column=1).border = _border()
        v = ws0.cell(row=i, column=2, value=val)
        v.font   = Font(name="Calibri", bold=True, size=12, color=MID_BLUE)
        v.fill   = _fill(WHITE)
        v.border = _border()
        v.alignment = _center()
        if kpi == "Accounts On Hold" and val > 0:
            v.font = Font(name="Calibri", bold=True, size=12, color=RED)
        if kpi == "Unvalidated Addresses" and val > 0:
            v.font = Font(name="Calibri", bold=True, size=12, color=ORANGE)

    ws0.column_dimensions["A"].width = 30
    ws0.column_dimensions["B"].width = 18

    # Now embed other sheets by reusing the export functions' logic
    # (we share the workbook rather than saving separately)

    conn.close()

    # Append all sheets from other reports
    masters   = export_customers_master.__wrapped__(wb) if hasattr(export_customers_master, "__wrapped__") else None
    # Re-generate sheets inside this wb
    _add_accounts_sheet(wb, c if False else get_connection().cursor())
    _add_credit_sheet(wb)
    _add_audit_sheet(wb)

    path = _get_output_path("Customer_Master_DASHBOARD", output_dir)
    wb.save(path)
    get_connection().close()

    return {
        "status": "SUCCESS",
        "file":   path,
        "sheets": [ws.title for ws in wb.worksheets],
    }


def _add_accounts_sheet(wb, c):
    ws = wb.create_sheet("Customer Accounts")
    headers = ["Account ID", "Account #", "Party Name", "Credit Limit ($)",
               "Open Balance ($)", "Lifecycle State", "Avg Days Pay", "On Hold"]
    _title_row(ws, "Customer Accounts", f"Generated: {datetime.now():%Y-%m-%d %H:%M}", len(headers))
    _apply_header_row(ws, headers, row=3, bg=MID_BLUE)
    rows = c.execute("""
        SELECT ca.cust_account_id, ca.account_number, hp.party_name,
               ca.credit_limit,
               COALESCE((SELECT SUM(ps.amount_remaining) FROM ar_payment_schedules ps
                          WHERE ps.cust_account_id=ca.cust_account_id AND ps.status='OP'),0),
               ca.lifecycle_state, ca.avg_days_to_pay, ca.on_hold
        FROM hz_cust_accounts ca
        JOIN hz_parties hp ON hp.party_id=ca.party_id
        ORDER BY ca.lifecycle_state
    """).fetchall()
    for i, row in enumerate(rows, 4):
        bg = LIGHT_GREY if i % 2 == 0 else WHITE
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = _body_font(); cell.border = _border(); cell.alignment = _left()
            cell.fill = _fill(bg)
        lc = LIFECYCLE_COLOURS.get(row[5], WHITE)
        ws.cell(row=i, column=6).fill = _fill(lc)
        ws.cell(row=i, column=6).font = Font(name="Calibri", bold=True, size=10,
                                              color=WHITE if row[5] in ("ACTIVE","INACTIVE") else "000000")
    _auto_width(ws)
    ws.freeze_panes = "A4"


def _add_credit_sheet(wb):
    from agents.credit import evaluate_credit
    conn = get_connection()
    ids = [r[0] for r in conn.execute(
        "SELECT cust_account_id FROM hz_cust_accounts WHERE status='A'"
    ).fetchall()]
    conn.close()
    evals = [evaluate_credit(i) for i in ids]
    evals.sort(key=lambda x: x.get("score", 0))

    ws = wb.create_sheet("Credit Analysis")
    headers = ["Account #", "Party Name", "Score", "Recommendation",
               "Current Limit ($)", "New Limit ($)", "Top Reason"]
    _title_row(ws, "Credit Analysis", f"Generated: {datetime.now():%Y-%m-%d %H:%M}", len(headers))
    _apply_header_row(ws, headers, row=3, bg=DARK_BLUE)
    RECC = {"INCREASE": GREEN, "NO_CHANGE": MID_BLUE, "DECREASE": ORANGE, "HOLD": RED}
    for i, ev in enumerate(evals, 4):
        if "error" in ev: continue
        bg = LIGHT_GREY if i % 2 == 0 else WHITE
        vals = [ev["account_number"], ev["party_name"], ev["score"],
                ev["recommendation"], ev["current_limit"], ev["new_limit"],
                ev["reasons"][0] if ev["reasons"] else ""]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = _body_font(); cell.border = _border(); cell.alignment = _left()
            cell.fill = _fill(bg)
        rec = ws.cell(row=i, column=4)
        rec.fill = _fill(RECC.get(ev["recommendation"], WHITE))
        rec.font = Font(name="Calibri", bold=True, size=10,
                        color=WHITE if ev["recommendation"] in ("INCREASE","HOLD") else "000000")
    _auto_width(ws)
    ws.freeze_panes = "A4"


def _add_audit_sheet(wb):
    conn = get_connection()
    rows = conn.execute("SELECT * FROM audit_log ORDER BY log_id DESC").fetchall()
    conn.close()
    ws = wb.create_sheet("Audit Log")
    headers = ["Log ID", "Workflow", "Entity Type", "Entity ID", "Action", "Performed At"]
    _title_row(ws, "AI Audit Log", f"Generated: {datetime.now():%Y-%m-%d %H:%M}  |  Entries: {len(rows)}", len(headers))
    _apply_header_row(ws, headers, row=3, bg=DARK_BLUE)
    WFC = {"DEDUPLICATION": MID_BLUE, "ADDRESS_VALIDATION": GREEN,
           "CREDIT_ADJUSTMENT": ORANGE, "CONTACT_MAINTENANCE": "7030A0",
           "RELATIONSHIP_MGMT": "00B0F0", "ARCHIVING": RED}
    for i, row in enumerate(rows, 4):
        r = dict(row)
        bg = LIGHT_GREY if i % 2 == 0 else WHITE
        vals = [r["log_id"], r["workflow"], r["entity_type"], r["entity_id"], r["action"], r["performed_at"]]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = _body_font(size=9); cell.border = _border(); cell.alignment = _left()
            cell.fill = _fill(bg)
        wfc = ws.cell(row=i, column=2)
        wfc.fill = _fill(WFC.get(r["workflow"], MID_BLUE))
        wfc.font = Font(name="Calibri", bold=True, size=9, color=WHITE)
    _auto_width(ws)
    ws.freeze_panes = "A4"
